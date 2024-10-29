# Tambahkan ini di awal file
print("Script dimulai")

import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

print("Imports selesai")


class Menu:
    MENU_ITEMS = {
        "Makanan Utama": {
            "Mujair + Nasi": 12000,
            "Lele + Nasi": 12000,
            "Belut Goreng + Nasi": 12000,
            "Ayam Goreng + Nasi": 12000,
            "Ayam Kampung + Nasi": 15000,
            "Bebek + Nasi": 15000,
            "Nasi Soto Lamongan": 12000,
            "Nasi Uduk": 10000
        },
        "Minuman": {
            "Es Jeruk": 4000,
            "Jeruk Hangat": 4000,
            "Es Teh": 3000,
            "Teh Hangat": 3000,
            "Kopi Hitam": 4000
        }
    }

    @classmethod
    def get_random_meal(cls):
        return random.choice(list(cls.MENU_ITEMS["Makanan Utama"].items()))

    @classmethod
    def get_random_drink(cls):
        return random.choice(list(cls.MENU_ITEMS["Minuman"].items()))
    
class WarungSotoFinancial:
    SOTO_PRICE = 12000
    GORENGAN_PRICE = 1000
    GAS_PER_TABUNG = 20000
    GAS_USAGE_DAYS = 4
    GAJI_KARYAWAN = 30000

    BUMBU = {
        # Bahan utama
        "Ayam": 30000,  # per kg
        "Tempe": 10000,  # per papan
        "Tahu": 15000,  # per kotak
        "Singkong": 8000,  # per kg (untuk keripik)
        "Tepung terigu": 12000,  # per kg
        "Ikan mujair": 25000,  # per kg
        "Ikan lele": 22000,  # per kg
        "Belut": 40000,  # per kg
        "Ayam kampung": 50000,  # per kg
        "Bebek": 45000,  # per kg

        "Bawang merah": 30000, "Bawang putih": 25000, "Jahe": 20000,
        "Kunyit": 15000, "Kemiri": 40000, "Merica": 100000,
        "Ketumbar": 80000, "Serai": 15000, "Daun jeruk": 5000,
        "Daun salam": 5000, "Garam": 10000, "Gula": 15000,
        "Penyedap rasa": 20000
    }

    BUMBU_USAGE = {
        # Bahan utama (penggunaan per hari)
        "Ayam": 3,  # 3 kg per hari
        "Tempe": 2,  # 2 papan per hari
        "Tahu": 1,  # 1 kotak per hari
        "Singkong": 1,  # 1 kg per hari
        "Tepung terigu": 0.5,  # 0.5 kg per hari
        "Ikan mujair": 2,  # 2 kg per hari
        "Ikan lele": 2,  # 2 kg per hari
        "Belut": 1,  # 1 kg per hari
        "Ayam kampung": 1,  # 1 kg per hari
        "Bebek": 1,  # 1 kg per hari

        # Bumbu-bumbu yang digunakan dalam jumlah kecil, umumnya dalam gram:
        "Bawang merah": 0.25,  # Menggunakan 0.25 kg (250 gram) bawang merah per hari
        "Bawang putih": 0.15,  # Menggunakan 0.15 kg (150 gram) bawang putih per hari
        "Jahe": 0.05,  # Menggunakan 0.05 kg (50 gram) jahe per hari
        "Kunyit": 0.03,  # Menggunakan 0.03 kg (30 gram) kunyit per hari
        "Kemiri": 0.1,  # Menggunakan 0.1 kg (100 gram) kemiri per hari
        "Merica": 0.02,  # Menggunakan 0.02 kg (20 gram) merica per hari
        "Ketumbar": 0.02,  # Menggunakan 0.02 kg (20 gram) ketumbar per hari
        "Serai": 0.1,  # Menggunakan 0.1 kg (100 gram) serai per hari
        "Daun jeruk": 0.3,  # Menggunakan 0.3 kg (300 gram) daun jeruk per hari
        "Daun salam": 0.2,  # Menggunakan 0.2 kg (200 gram) daun salam per hari
        "Garam": 0.1,  # Menggunakan 0.1 kg (100 gram) garam per hari
        "Gula": 0.05,  # Menggunakan 0.05 kg (50 gram) gula per hari
        "Penyedap rasa": 0.02,  # Menggunakan 0.02 kg (20 gram) penyedap rasa per hari
        "Santan": 0.5,  # Menggunakan 0.5 kg santan per hari
        "Daun pandan": 0.05,  # Menggunakan 0.05 kg daun pandan per hari
        "Lengkuas": 0.05,  # Menggunakan 0.05 kg lengkuas per hari
        "Jeruk": 1,  # Menggunakan 1 kg jeruk per hari
        "Teh": 0.1,  # Menggunakan 0.1 kg teh per hari
        "Kopi": 0.2  # Menggunakan 0.2 kg kopi per hari
    }

    @staticmethod
    def get_indonesian_day_name(date):
        day_names = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
        return day_names[date.weekday()]

    @staticmethod
    def round_to_nearest(x, base=500):
        return base * round(x/base)

    @classmethod
    def calculate_spice_cost(cls):
        return sum(cls.BUMBU[spice] * cls.BUMBU_USAGE[spice] for spice in cls.BUMBU)

    @classmethod
    def generate_sales_data(cls, start_date, end_date):
        print(f"Generating sales data from {start_date} to {end_date}")
        data = []
        current_date = start_date
        while current_date <= end_date:
            print(f"Processing date: {current_date}")
            if current_date.weekday() != 6:  # Not Sunday
                print("  Not Sunday, generating sales")
                
                # Generate a random total sales amount within the range
                total_sales = cls.round_to_nearest(random.randint(500000, 720000), 50000)
                print(f"  Total sales: {total_sales}")

                # Generate random meals and drinks
                meals = {}
                drinks = {}
                remaining_sales = total_sales

                while remaining_sales > 0:
                    meal, price = Menu.get_random_meal()
                    if price <= remaining_sales:
                        meals[meal] = meals.get(meal, 0) + 1
                        remaining_sales -= price

                    drink, price = Menu.get_random_drink()
                    if price <= remaining_sales:
                        drinks[drink] = drinks.get(drink, 0) + 1
                        remaining_sales -= price

                    if remaining_sales < min(Menu.MENU_ITEMS["Minuman"].values()):
                        break

                # Calculate actual total sales based on meals and drinks sold
                total_sales = sum(Menu.MENU_ITEMS["Makanan Utama"][meal] * portions for meal, portions in meals.items()) + \
                            sum(Menu.MENU_ITEMS["Minuman"][drink] * portions for drink, portions in drinks.items())

                print(f"  Meals sold: {meals}")
                print(f"  Drinks sold: {drinks}")
                print(f"  Adjusted total sales: {total_sales}")

                print("  Calculating costs")
                spice_cost = cls.calculate_spice_cost()
                gas_cost = cls.GAS_PER_TABUNG / cls.GAS_USAGE_DAYS
                expenses = spice_cost + gas_cost + cls.GAJI_KARYAWAN
                profit = total_sales - expenses

                print("  Appending data for the day")
                data.append({
                    'date': current_date.strftime('%d %B %Y'),
                    'day': cls.get_indonesian_day_name(current_date),
                    'meals': meals,
                    'drinks': drinks,
                    'total_sales': total_sales,
                    'spice_cost': spice_cost,
                    'gas_cost': gas_cost,
                    'salary': cls.GAJI_KARYAWAN,
                    'expenses': expenses,
                    'profit': profit
                })
            else:
                print("  Sunday, shop is closed")
                data.append({
                    'date': current_date.strftime('%d %B %Y'),
                    'day': cls.get_indonesian_day_name(current_date),
                    'meals': 'Tutup',
                    'drinks': 'Tutup',
                    'total_sales': 'Tutup',
                    'spice_cost': 'Tutup',
                    'gas_cost': 'Tutup',
                    'salary': 'Tutup',
                    'expenses': 'Tutup',
                    'profit': 'Tutup'
                })
            print(f"  Moving to next date")
            current_date += timedelta(days=1)
        print(f"Generated {len(data)} days of sales data")
        return data

    @classmethod
    def generate_monthly_expenses(cls, work_days):
        print(f"Generating monthly expenses for {work_days} work days")
        return {
            "Bumbu": cls.round_to_nearest(work_days * cls.calculate_spice_cost()),
            "Gas": cls.round_to_nearest((work_days / cls.GAS_USAGE_DAYS) * cls.GAS_PER_TABUNG),
            "Air": cls.round_to_nearest(random.randint(200000, 300000)),
            "Listrik": cls.round_to_nearest(random.randint(200000, 400000)),
            "Gaji karyawan": work_days * cls.GAJI_KARYAWAN,
        }

    @staticmethod
    def generate_weekly_summary(sales_data):
        print("Generating weekly summary")
        weekly_summaries = []
        current_week = []
        totals = {'meals': {}, 'drinks': {}, 'sales': 0, 'expenses': 0, 'profit': 0, 'spice_cost': 0, 'total_soto': 0, 'total_gorengan': 0}

        for day in sales_data:
            if day['day'] != 'Minggu' and day['total_sales'] != 'Tutup':
                current_week.append(day)
                for meal, count in day['meals'].items():
                    if 'Soto' in meal:
                        totals['total_soto'] += count
                    elif 'Gorengan' in meal:
                        totals['total_gorengan'] += count
                    totals['meals'][meal] = totals['meals'].get(meal, 0) + count
                for drink, count in day['drinks'].items():
                    totals['drinks'][drink] = totals['drinks'].get(drink, 0) + count
                totals['sales'] += day['total_sales']
                totals['expenses'] += day['expenses']
                totals['profit'] += day['profit']
                totals['spice_cost'] += day['spice_cost']
            elif day['day'] == 'Minggu' or day == sales_data[-1]:
                weekly_summaries.append({
                    'start_date': current_week[0]['date'],
                    'end_date': day['date'],
                    'total_soto': totals['total_soto'],
                    'total_gorengan': totals['total_gorengan'],
                    'total_meals': totals['meals'],
                    'total_drinks': totals['drinks'],
                    'total_sales': totals['sales'],
                    'total_expenses': totals['expenses'],
                    'total_profit': totals['profit'],
                    'weekly_spice_cost': totals['spice_cost']
                })
                totals = {'meals': {}, 'drinks': {}, 'sales': 0, 'expenses': 0, 'profit': 0, 'spice_cost': 0, 'total_soto': 0, 'total_gorengan': 0}
                current_week = []

        print(f"Generated {len(weekly_summaries)} weekly summaries")
        return weekly_summaries

class ExcelExporter:
    @staticmethod
    def export_to_excel(sales_data, monthly_expenses, weekly_summaries, filename="Laporan_Keuangan_Warung.xlsx"):
        print(f"Exporting data to Excel: {filename}")
        wb = Workbook()
        ExcelExporter._create_daily_sales_sheet(wb, sales_data)
        ExcelExporter._create_monthly_expenses_sheet(wb, monthly_expenses)
        ExcelExporter._create_weekly_summary_sheet(wb, weekly_summaries)
        ExcelExporter._create_monthly_summary_sheet(wb, sales_data, monthly_expenses)
        wb.save(filename)
        print(f"Excel file saved: {filename}")

    @staticmethod
    def _create_daily_sales_sheet(wb, sales_data):
        print("Creating daily sales sheet")
        ws = wb.active
        ws.title = "Penjualan Harian"
        headers = ["Tanggal", "Hari", "Makanan (Porsi)", "Minuman (Porsi)", "Total Penjualan (Rp)", 
                "Biaya Bumbu (Rp)", "Biaya Gas (Rp)", "Gaji (Rp)", "Pengeluaran (Rp)", "Laba (Rp)"]
        ws.append(headers)

        for day in sales_data:
            if day['total_sales'] != 'Tutup':
                meals = ", ".join([f"{meal}: {count}" for meal, count in day['meals'].items()])
                drinks = ", ".join([f"{drink}: {count}" for drink, count in day['drinks'].items()])
                ws.append([
                    day['date'], day['day'], meals, drinks,
                    "{:,.0f}".format(day['total_sales']), "{:,.0f}".format(day['spice_cost']),
                    "{:,.0f}".format(day['gas_cost']), "{:,.0f}".format(day['salary']),
                    "{:,.0f}".format(day['expenses']), "{:,.0f}".format(day['profit'])
                ])
            else:
                ws.append([day['date'], day['day']] + ['Tutup'] * 8)

        ExcelExporter._format_sheet(ws, headers)

    @staticmethod
    def _create_monthly_expenses_sheet(wb, monthly_expenses):
        print("Creating monthly expenses sheet")
        ws = wb.create_sheet(title="Pengeluaran Bulanan")
        ws.append(["Jenis Pengeluaran", "Jumlah (Rp)"])
        for expense, amount in monthly_expenses.items():
            ws.append([expense, "{:,.0f}".format(amount)])

        ws.append(["", ""])
        ws.append(["Detail Pengeluaran Bumbu (per hari)", ""])
        for spice, price in WarungSotoFinancial.BUMBU.items():
            daily_cost = price * WarungSotoFinancial.BUMBU_USAGE[spice]
            ws.append([spice, "{:,.0f}".format(daily_cost)])

        ExcelExporter._format_sheet(ws, ["Jenis Pengeluaran", "Jumlah (Rp)"])

    @staticmethod
    def _create_weekly_summary_sheet(wb, weekly_summaries):
        print("Creating weekly summary sheet")
        ws = wb.create_sheet(title="Ringkasan Mingguan")
        headers = ["Tanggal Mulai", "Tanggal Akhir", "Total Soto (Porsi)", "Total Gorengan (Porsi)",
                "Total Penjualan (Rp)", "Total Pengeluaran (Rp)", "Total Laba (Rp)", "Biaya Bumbu Mingguan (Rp)"]
        ws.append(headers)

        for week in weekly_summaries:
            ws.append([
                week['start_date'], week['end_date'], week['total_soto'], week['total_gorengan'],
                "{:,.0f}".format(week['total_sales']), "{:,.0f}".format(week['total_expenses']),
                "{:,.0f}".format(week['total_profit']), "{:,.0f}".format(week['weekly_spice_cost'])
            ])

        ExcelExporter._format_sheet(ws, headers)


    @staticmethod
    def _create_monthly_summary_sheet(wb, sales_data, monthly_expenses):
        print("Creating monthly summary sheet")
        ws = wb.create_sheet(title="Ringkasan Bulanan")
        total_monthly_sales = sum(day['total_sales'] for day in sales_data if day['total_sales'] != 'Tutup')
        total_monthly_expenses = sum(monthly_expenses.values())
        monthly_profit = total_monthly_sales - total_monthly_expenses

        total_soto_portions = 0
        total_gorengan_portions = 0
        for day in sales_data:
            if day['total_sales'] != 'Tutup':
                for meal, count in day['meals'].items():
                    if 'Soto' in meal:
                        total_soto_portions += count
                    elif 'Gorengan' in meal:
                        total_gorengan_portions += count

        summary_data = [
            ["Ringkasan Bulanan", "Jumlah"],
            ["Total Penjualan (Rp)", "{:,.0f}".format(total_monthly_sales)],
            ["Total Pengeluaran (Rp)", "{:,.0f}".format(total_monthly_expenses)],
            ["Laba/Rugi Bulanan (Rp)", "{:,.0f}".format(monthly_profit)],
            ["Total Porsi Soto", total_soto_portions],
            ["Total Porsi Gorengan", total_gorengan_portions]
        ]

        for row in summary_data:
            ws.append(row)

        ExcelExporter._format_sheet(ws, ["Ringkasan Bulanan", "Jumlah"])

    @staticmethod
    def _format_sheet(ws, headers):
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

def main():
    print("Fungsi main() dimulai")
    try:
        start_date = datetime(2024, 7, 29)
        end_date = datetime(2024, 8, 29)
        print(f"Date range: {start_date} to {end_date}")

        print("Generating sales data...")
        sales_data = WarungSotoFinancial.generate_sales_data(start_date, end_date)
        print("Sales data generated")

        print("Calculating work days...")
        work_days = sum(1 for day in sales_data if day['total_sales'] != 'Tutup')
        print(f"Work days: {work_days}")

        print("Generating monthly expenses...")
        monthly_expenses = WarungSotoFinancial.generate_monthly_expenses(work_days)
        print("Monthly expenses generated")

        print("Generating weekly summaries...")
        weekly_summaries = WarungSotoFinancial.generate_weekly_summary(sales_data)
        print("Weekly summaries generated")

        print("Exporting to Excel...")
        ExcelExporter.export_to_excel(sales_data, monthly_expenses, weekly_summaries)
        print("Excel export completed")

        print("Program finished successfully")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

print("Sebelum if __name__ == '__main__'")
if __name__ == "__main__":
    print("Memanggil main()")
    main()
print("Setelah pemanggilan main()")
