import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta

# Define the expense details
expenses = {
    "Ayam": 90000,
    "Tempe": 20000,
    "Tahu": 15000,
    "Singkong": 8000,
    "Tepung terigu": 6000,
    "Ikan mujair": 50000,
    "Ikan lele": 44000,
    "Belut": 40000,
    "Ayam kampung": 50000,
    "Bebek": 45000,
    "Bawang merah": 7500,
    "Bawang putih": 3750,
    "Jahe": 1000,
    "Kunyit": 450,
    "Kemiri": 4000,
    "Merica": 2000,
    "Ketumbar": 1600,
    "Serai": 1500,
    "Daun jeruk": 1500,
    "Daun salam": 1000,
    "Garam": 1000,
    "Gula": 750,
    "Penyedap rasa": 400
}

# Set the start and end dates
start_date = datetime(2024, 7, 29)
end_date = datetime(2024, 8, 29)

# Create a new workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Pengeluaran Bumbu"

# Set the column widths
worksheet.column_dimensions['A'].width = 15
worksheet.column_dimensions['B'].width = 15
worksheet.column_dimensions['C'].width = 15
worksheet.column_dimensions['D'].width = 15
worksheet.column_dimensions['E'].width = 15

# Add the header row
header_row = ['Tanggal', 'Hari', 'Pengeluaran', 'Total', '']
for col, value in enumerate(header_row, 1):
    cell = worksheet.cell(row=1, column=col, value=value)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Add the expense data to the worksheet
row = 2
current_date = start_date
while current_date <= end_date:
    if current_date.weekday() != 6:  # Skip Sundays
        worksheet.cell(row=row, column=1, value=current_date.strftime('%d/%m/%Y'))
        worksheet.cell(row=row, column=2, value=current_date.strftime('%A'))

        total_expense = 0
        for item, cost in expenses.items():
            worksheet.cell(row=row, column=3, value=item)
            worksheet.cell(row=row, column=4, value=cost)
            total_expense += cost
            row += 1

        worksheet.cell(row=row, column=3, value='Total')
        worksheet.cell(row=row, column=4, value=total_expense)
        row += 2

    current_date += timedelta(days=1)

# Save the workbook
workbook.save('pengeluaran_bumbu.xlsx')
print("Excel file created successfully!")